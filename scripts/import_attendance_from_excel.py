import argparse
import datetime as dt
import json
import os
import sys
import unicodedata
from collections import defaultdict
from typing import Any

import requests
from openpyxl import load_workbook


TARGET_SHEETS = ["M10", "M11", "M12", "M13"]


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\n", " ").strip().lower()
    text = "".join(
        ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch)
    )
    return " ".join(text.split())


def clean_text(value: Any) -> str | None:
    text = normalize_text(value)
    return text or None


def digits_only(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value))
    digits = "".join(ch for ch in str(value) if ch.isdigit())
    return digits or None


def parse_date(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()

    text = str(value).strip().replace("//", "/")
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return dt.datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def is_present(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    text = normalize_text(value)
    return text in {"1", "si", "s", "x", "ok", "presente", "p"}


def chunked(items: list[dict[str, Any]], size: int) -> list[list[dict[str, Any]]]:
    return [items[i:i + size] for i in range(0, len(items), size)]


class SupabaseRest:
    def __init__(self, base_url: str, api_key: str):
        self.base_url = base_url.rstrip("/")
        self.api_key = api_key
        self.base_headers = {
            "apikey": api_key,
            "Authorization": f"Bearer {api_key}",
        }

    def get(self, table: str, params: dict[str, Any]) -> list[dict[str, Any]]:
        response = requests.get(
            f"{self.base_url}/rest/v1/{table}",
            headers=self.base_headers,
            params=params,
            timeout=60,
        )
        response.raise_for_status()
        return response.json()

    def post(self, table: str, payload: list[dict[str, Any]] | dict[str, Any], prefer: str) -> list[dict[str, Any]]:
        headers = {
            **self.base_headers,
            "Content-Type": "application/json",
            "Prefer": prefer,
        }
        response = requests.post(
            f"{self.base_url}/rest/v1/{table}",
            headers=headers,
            data=json.dumps(payload),
            timeout=60,
        )
        response.raise_for_status()
        return response.json() if response.text else []


def get_divisions(api: SupabaseRest) -> dict[str, str]:
    rows = api.get(
        "divisions",
        {
            "select": "id,name",
            "name": f"in.({','.join(TARGET_SHEETS)})",
        },
    )
    return {row["name"]: row["id"] for row in rows}


def get_players(api: SupabaseRest, division_ids: list[str]) -> list[dict[str, Any]]:
    return api.get(
        "players",
        {
            "select": "id,first_name,last_name,dni,birth_date,division_id",
            "division_id": f"in.({','.join(division_ids)})",
        },
    )


def get_sessions(api: SupabaseRest, division_ids: list[str], dates: list[str]) -> list[dict[str, Any]]:
    quoted_dates = ",".join(f'"{d}"' for d in dates)
    return api.get(
        "training_sessions",
        {
            "select": "id,division_id,session_date",
            "division_id": f"in.({','.join(division_ids)})",
            "session_date": f"in.({quoted_dates})",
        },
    )


def build_player_indexes(players: list[dict[str, Any]]) -> tuple[dict[tuple[str, str], str], dict[tuple[str, str, str], str], dict[tuple[str, str], list[str]], dict[tuple[str, str, str], list[str]]]:
    by_dni: dict[tuple[str, str], str] = {}
    by_name_birth: dict[tuple[str, str, str], str] = {}
    ambiguous_name: dict[tuple[str, str], list[str]] = defaultdict(list)
    ambiguous_name_birth: dict[tuple[str, str, str], list[str]] = defaultdict(list)

    for player in players:
        division_id = player["division_id"]
        first_name = clean_text(player.get("first_name")) or ""
        last_name = clean_text(player.get("last_name")) or ""
        name_key = f"{first_name}|{last_name}"
        birth_date = player.get("birth_date")
        dni = digits_only(player.get("dni"))

        ambiguous_name[(division_id, name_key)].append(player["id"])
        if birth_date:
            ambiguous_name_birth[(division_id, name_key, birth_date)].append(player["id"])
        if dni:
            by_dni[(division_id, dni)] = player["id"]
        if birth_date:
            by_name_birth[(division_id, name_key, birth_date)] = player["id"]

    return by_dni, by_name_birth, ambiguous_name, ambiguous_name_birth


def find_player_id(
    division_id: str,
    dni: str | None,
    first_name: str | None,
    last_name: str | None,
    birth_date: str | None,
    by_dni: dict[tuple[str, str], str],
    by_name_birth: dict[tuple[str, str, str], str],
    ambiguous_name: dict[tuple[str, str], list[str]],
    ambiguous_name_birth: dict[tuple[str, str, str], list[str]],
) -> tuple[str | None, str]:
    if dni:
        player_id = by_dni.get((division_id, dni))
        if player_id:
            return player_id, "dni"

    first = first_name or ""
    last = last_name or ""
    name_key = f"{first}|{last}"

    if birth_date:
        birth_matches = ambiguous_name_birth.get((division_id, name_key, birth_date), [])
        if len(birth_matches) == 1:
            return birth_matches[0], "name+birth"
        player_id = by_name_birth.get((division_id, name_key, birth_date))
        if player_id:
            return player_id, "name+birth"

    name_matches = ambiguous_name.get((division_id, name_key), [])
    if len(name_matches) == 1:
        return name_matches[0], "name"
    if len(name_matches) > 1:
        return None, "ambiguous-name"

    return None, "not-found"


def build_import_plan(workbook_path: str, divisions: dict[str, str], players: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    wb = load_workbook(workbook_path, data_only=True)
    by_dni, by_name_birth, ambiguous_name, ambiguous_name_birth = build_player_indexes(players)

    sessions_needed: dict[tuple[str, str], dict[str, Any]] = {}
    attendance_records: list[dict[str, Any]] = []
    unmatched: list[dict[str, Any]] = []
    stats: dict[str, Any] = {
        "rows_by_sheet": defaultdict(int),
        "matched_by_sheet": defaultdict(int),
        "unmatched_by_sheet": defaultdict(int),
        "present_cells": 0,
        "records_total": 0,
        "dates": set(),
    }

    for sheet_name in wb.sheetnames:
        canonical_sheet = normalize_text(sheet_name).upper()
        if canonical_sheet not in TARGET_SHEETS:
            continue

        ws = wb[sheet_name]
        header_values = [cell.value for cell in ws[1]]
        date_columns: list[tuple[int, str]] = []
        for idx, header in enumerate(header_values):
            parsed = parse_date(header)
            if parsed:
                date_columns.append((idx, parsed))
                stats["dates"].add(parsed)

        division_id = divisions[canonical_sheet]

        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            first_name = clean_text(row[4] if len(row) > 4 else None)
            last_name = clean_text(row[5] if len(row) > 5 else None)
            if canonical_sheet == "M11":
                first_name = clean_text(row[3] if len(row) > 3 else None)
                last_name = clean_text(row[4] if len(row) > 4 else None)

            if not first_name and not last_name:
                continue

            stats["rows_by_sheet"][canonical_sheet] += 1

            dni = digits_only(row[2] if len(row) > 2 else None)
            birth_idx = 3 if canonical_sheet != "M11" else None
            birth_date = parse_date(row[birth_idx]) if birth_idx is not None and len(row) > birth_idx else None

            player_id, match_type = find_player_id(
                division_id,
                dni,
                first_name,
                last_name,
                birth_date,
                by_dni,
                by_name_birth,
                ambiguous_name,
                ambiguous_name_birth,
            )

            if not player_id:
                stats["unmatched_by_sheet"][canonical_sheet] += 1
                unmatched.append(
                    {
                        "sheet": canonical_sheet,
                        "row": row_index,
                        "first_name": first_name,
                        "last_name": last_name,
                        "dni": dni,
                        "birth_date": birth_date,
                        "reason": match_type,
                    }
                )
                continue

            stats["matched_by_sheet"][canonical_sheet] += 1

            for col_idx, session_date in date_columns:
                present = is_present(row[col_idx] if len(row) > col_idx else None)
                if present:
                    stats["present_cells"] += 1
                sessions_needed[(division_id, session_date)] = {
                    "division_id": division_id,
                    "session_date": session_date,
                    "session_type": "sabado",
                }
                attendance_records.append(
                    {
                        "division_id": division_id,
                        "session_date": session_date,
                        "player_id": player_id,
                        "present": present,
                    }
                )

    stats["records_total"] = len(attendance_records)
    stats["dates"] = sorted(stats["dates"])
    return list(sessions_needed.values()), attendance_records, unmatched, stats


def ensure_sessions(api: SupabaseRest, sessions_needed: list[dict[str, Any]]) -> dict[tuple[str, str], str]:
    if sessions_needed:
        for batch in chunked(sessions_needed, 100):
            api.post("training_sessions", batch, "resolution=merge-duplicates,return=representation")

    division_ids = sorted({row["division_id"] for row in sessions_needed})
    dates = sorted({row["session_date"] for row in sessions_needed})
    existing = get_sessions(api, division_ids, dates)
    return {(row["division_id"], row["session_date"]): row["id"] for row in existing}


def insert_attendance(api: SupabaseRest, records: list[dict[str, Any]], session_map: dict[tuple[str, str], str]) -> None:
    payload: list[dict[str, Any]] = []
    for row in records:
        session_id = session_map[(row["division_id"], row["session_date"])]
        payload.append(
            {
                "session_id": session_id,
                "player_id": row["player_id"],
                "present": row["present"],
            }
        )

    for batch in chunked(payload, 500):
        api.post("attendance_records", batch, "resolution=merge-duplicates,return=minimal")


def summarize_sessions(api: SupabaseRest, session_ids: list[str]) -> list[dict[str, Any]]:
    quoted_ids = ",".join(f'"{sid}"' for sid in session_ids)
    sessions = api.get(
        "training_sessions",
        {
            "select": "id,session_date,division_id,divisions(name)",
            "id": f"in.({quoted_ids})",
            "order": "session_date.asc",
        },
    )
    attendance = api.get(
        "attendance_records",
        {
            "select": "session_id,present",
            "session_id": f"in.({quoted_ids})",
        },
    )

    counts: dict[str, dict[str, int]] = defaultdict(lambda: {"present": 0, "total": 0})
    for row in attendance:
        counts[row["session_id"]]["total"] += 1
        if row["present"]:
            counts[row["session_id"]]["present"] += 1

    summary: list[dict[str, Any]] = []
    for session in sessions:
        division = session.get("divisions")
        division_name = division["name"] if isinstance(division, dict) else division[0]["name"]
        summary.append(
            {
                "division": division_name,
                "date": session["session_date"],
                "present": counts[session["id"]]["present"],
                "total": counts[session["id"]]["total"],
            }
        )
    return summary


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", required=True, help="Ruta al Excel de presentismos")
    parser.add_argument("--execute", action="store_true", help="Inserta en Supabase")
    args = parser.parse_args()

    base_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    api_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not base_url or not api_key:
        print("Faltan NEXT_PUBLIC_SUPABASE_URL o SUPABASE_SERVICE_ROLE_KEY.", file=sys.stderr)
        return 1

    api = SupabaseRest(base_url, api_key)
    divisions = get_divisions(api)
    missing_divisions = [sheet for sheet in TARGET_SHEETS if sheet not in divisions]
    if missing_divisions:
        print(f"Faltan divisiones en Supabase: {', '.join(missing_divisions)}", file=sys.stderr)
        return 1

    players = get_players(api, list(divisions.values()))
    sessions_needed, attendance_records, unmatched, stats = build_import_plan(args.file, divisions, players)

    print("Vista previa de presentismos:")
    for sheet in TARGET_SHEETS:
        print(
            f"- {sheet}: filas={stats['rows_by_sheet'][sheet]}, "
            f"matcheadas={stats['matched_by_sheet'][sheet]}, "
            f"sin_match={stats['unmatched_by_sheet'][sheet]}"
        )
    print(f"- Fechas detectadas: {', '.join(stats['dates'])}")
    print(f"- Sesiones a usar/crear: {len(sessions_needed)}")
    print(f"- Registros de asistencia: {stats['records_total']}")
    print(f"- Presentes marcados: {stats['present_cells']}")

    if unmatched:
        print("")
        print("Filas sin match:")
        for row in unmatched[:20]:
            print(
                f"- {row['sheet']} fila {row['row']}: "
                f"{(row['first_name'] or '').title()} {(row['last_name'] or '').title()} "
                f"(dni={row['dni']}, nac={row['birth_date']}) motivo={row['reason']}"
            )
        if len(unmatched) > 20:
            print(f"- ... y {len(unmatched) - 20} mas")

    if not args.execute:
        print("")
        print("Modo vista previa. Usa --execute para importar.")
        return 0

    session_map = ensure_sessions(api, sessions_needed)
    insert_attendance(api, attendance_records, session_map)

    print("")
    print("Resumen en Supabase:")
    for row in summarize_sessions(api, list(session_map.values())):
        print(f"- {row['division']} {row['date']}: {row['present']}/{row['total']}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
