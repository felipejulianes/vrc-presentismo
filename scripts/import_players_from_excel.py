import argparse
import datetime as dt
import json
import os
import sys
import unicodedata
from typing import Any

import requests
from openpyxl import load_workbook


TARGET_SHEETS = ["M10", "M11", "M12", "M13", "M14"]


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\n", " ").strip().lower()
    text = "".join(
        ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch)
    )
    return " ".join(text.split())


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return " ".join(text.split())


def digits_only(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value))
    text = "".join(ch for ch in str(value) if ch.isdigit())
    return text or None


def parse_date(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()

    text = str(value).strip()
    if not text:
        return None

    text = text.replace("//", "/")
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return dt.datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"Fecha invalida: {value!r}")


def parse_active(value: Any) -> bool:
    if value in (None, ""):
        return True
    text = normalize_header(value)
    if text in {"si", "s", "yes", "true", "1", "activo"}:
        return True
    if text in {"no", "n", "false", "0", "inactivo"}:
        return False
    return True


def chunked(items: list[dict[str, Any]], size: int) -> list[list[dict[str, Any]]]:
    return [items[i:i + size] for i in range(0, len(items), size)]


def get_division_map(base_url: str, api_key: str) -> dict[str, str]:
    response = requests.get(
        f"{base_url}/rest/v1/divisions",
        headers={
            "apikey": api_key,
            "Authorization": f"Bearer {api_key}",
        },
        params={
            "select": "id,name",
            "name": f"in.({','.join(TARGET_SHEETS)})",
        },
        timeout=30,
    )
    response.raise_for_status()
    rows = response.json()
    return {row["name"]: row["id"] for row in rows}


def build_header_map(headers: list[str]) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for idx, header in enumerate(headers):
        if header and header not in header_map:
            header_map[header] = idx
    return header_map


def build_payloads(workbook_path: str, division_map: dict[str, str]) -> tuple[list[dict[str, Any]], dict[str, int]]:
    wb = load_workbook(workbook_path, data_only=True)
    today = dt.date.today().isoformat()
    payloads: list[dict[str, Any]] = []
    counts: dict[str, int] = {}
    seen_dni: dict[str, int] = {}
    duplicate_dnis: set[str] = set()

    for sheet_name in TARGET_SHEETS:
        ws = wb[sheet_name]
        headers = [normalize_header(cell.value) for cell in ws[1]]
        header_map = build_header_map(headers)

        counts[sheet_name] = 0

        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            first_name = clean_text(row[header_map["nombre"]]) if "nombre" in header_map else None
            last_name = clean_text(row[header_map["apellido"]]) if "apellido" in header_map else None

            if not first_name and not last_name:
                continue

            active_value = row[header_map["activo"]] if "activo" in header_map else None
            dni = digits_only(row[header_map["dni"]]) if "dni" in header_map else None
            birth_date = (
                parse_date(row[header_map["fecha de nacimiento"]])
                if "fecha de nacimiento" in header_map
                else None
            )
            parent_phone = (
                digits_only(row[header_map["telefono referente"]])
                if "telefono referente" in header_map
                else None
            )
            parent_name = (
                clean_text(row[header_map["nombre referente"]])
                if "nombre referente" in header_map
                else None
            )
            is_active = parse_active(active_value)

            if dni:
                if dni in seen_dni:
                    duplicate_dnis.add(dni)
                else:
                    seen_dni[dni] = len(payloads)

            if dni in duplicate_dnis:
                first_index = seen_dni.get(dni)
                if first_index is not None:
                    payloads[first_index]["dni"] = None
                dni = None

            payloads.append(
                {
                    "first_name": first_name or "",
                    "last_name": last_name or "",
                    "dni": dni,
                    "birth_date": birth_date,
                    "parent_phone": parent_phone,
                    "parent_name": parent_name,
                    "division_id": division_map[sheet_name],
                    "active": True,
                    "inactivo": not is_active,
                    "fecha_alta": today,
                }
            )
            counts[sheet_name] += 1

    return payloads, counts


def insert_players(base_url: str, api_key: str, payloads: list[dict[str, Any]]) -> None:
    headers = {
        "apikey": api_key,
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }

    for batch in chunked(payloads, 100):
        response = requests.post(
            f"{base_url}/rest/v1/players",
            headers=headers,
            data=json.dumps(batch),
            timeout=60,
        )
        response.raise_for_status()


def fetch_imported_counts(base_url: str, api_key: str) -> list[dict[str, Any]]:
    response = requests.get(
        f"{base_url}/rest/v1/divisions",
        headers={
            "apikey": api_key,
            "Authorization": f"Bearer {api_key}",
        },
        params={
            "select": "name,players(count)",
            "name": f"in.({','.join(TARGET_SHEETS)})",
            "order": "sort_order.asc",
        },
        timeout=30,
    )
    response.raise_for_status()
    return response.json()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", required=True, help="Ruta al archivo Excel")
    parser.add_argument("--execute", action="store_true", help="Inserta en Supabase")
    args = parser.parse_args()

    base_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    api_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not base_url or not api_key:
        print("Faltan NEXT_PUBLIC_SUPABASE_URL o SUPABASE_SERVICE_ROLE_KEY.", file=sys.stderr)
        return 1

    division_map = get_division_map(base_url, api_key)
    missing = [sheet for sheet in TARGET_SHEETS if sheet not in division_map]
    if missing:
        print(f"Faltan divisiones en Supabase: {', '.join(missing)}", file=sys.stderr)
        return 1

    payloads, counts = build_payloads(args.file, division_map)

    print("Vista previa de importacion:")
    for sheet in TARGET_SHEETS:
        print(f"- {sheet}: {counts.get(sheet, 0)} jugadores")
    print(f"- Total: {len(payloads)} jugadores")

    if not args.execute:
        print("")
        print("Modo vista previa. Usa --execute para insertar.")
        return 0

    insert_players(base_url, api_key, payloads)

    print("")
    print("Conteos en Supabase despues de importar:")
    for row in fetch_imported_counts(base_url, api_key):
        count_rows = row.get("players") or []
        count = count_rows[0]["count"] if count_rows else 0
        print(f"- {row['name']}: {count}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
