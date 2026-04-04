import argparse
import datetime as dt
import json
import os
import sys
import unicodedata
from collections import defaultdict
from typing import Any

import requests
from openpyxl import load_workbook


TARGET_DIVISIONS = ["M6", "M7", "M8", "M9"]


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\n", " ").strip().lower()
    text = "".join(
        ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch)
    )
    return " ".join(text.split())


def clean_text(value: Any) -> str | None:
    text = "" if value is None else str(value).strip()
    if not text:
        return None
    return " ".join(text.split())


def digits_only(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value))
    digits = "".join(ch for ch in str(value) if ch.isdigit())
    return digits or None


def parse_date(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = str(value).strip().replace("//", "/")
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return dt.datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def parse_active(value: Any) -> bool:
    if value in (None, ""):
        return True
    text = normalize_text(value)
    if text in {"si", "s", "yes", "true", "1", "activo"}:
        return True
    if text in {"no", "n", "false", "0", "inactivo"}:
        return False
    return True


def is_present(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    text = normalize_text(value)
    return text in {"1", "si", "s", "x", "ok", "presente", "p"}


def chunked(items: list[dict[str, Any]], size: int) -> list[list[dict[str, Any]]]:
    return [items[i:i + size] for i in range(0, len(items), size)]


class SupabaseRest:
    def __init__(self, base_url: str, api_key: str):
        self.base_url = base_url.rstrip("/")
        self.api_key = api_key
        self.headers = {
            "apikey": api_key,
            "Authorization": f"Bearer {api_key}",
        }

    def get(self, table: str, params: dict[str, Any]) -> list[dict[str, Any]]:
        response = requests.get(
            f"{self.base_url}/rest/v1/{table}",
            headers=self.headers,
            params=params,
            timeout=60,
        )
        response.raise_for_status()
        return response.json()

    def post(self, table: str, payload: list[dict[str, Any]] | dict[str, Any], prefer: str) -> list[dict[str, Any]]:
        response = requests.post(
            f"{self.base_url}/rest/v1/{table}",
            headers={
                **self.headers,
                "Content-Type": "application/json",
                "Prefer": prefer,
            },
            data=json.dumps(payload),
            timeout=60,
        )
        response.raise_for_status()
        return response.json() if response.text else []


def get_divisions(api: SupabaseRest) -> dict[str, str]:
    rows = api.get(
        "divisions",
        {
            "select": "id,name",
            "name": f"in.({','.join(TARGET_DIVISIONS)})",
        },
    )
    return {row["name"]: row["id"] for row in rows}


def build_plan(workbook_path: str, divisions: dict[str, str]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    wb = load_workbook(workbook_path, data_only=True)
    today = dt.date.today().isoformat()
    players: list[dict[str, Any]] = []
    attendance_rows: list[dict[str, Any]] = []
    stats: dict[str, Any] = {
        "players_by_division": defaultdict(int),
        "present_cells": 0,
        "dates": set(),
    }
    seen_dni: set[str] = set()
    duplicate_dni: set[str] = set()

    for ws in wb.worksheets:
        header = [cell.value for cell in ws[1]]
        date_columns: list[tuple[int, str]] = []
        for idx, value in enumerate(header):
            parsed = parse_date(value)
            if parsed:
                date_columns.append((idx, parsed))
                stats["dates"].add(parsed)

        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            division_name = clean_text(row[0] if len(row) > 0 else None)
            if division_name not in TARGET_DIVISIONS:
                continue

            first_name = clean_text(row[5] if len(row) > 5 else None)
            last_name = clean_text(row[6] if len(row) > 6 else None)
            if not first_name and not last_name:
                continue

            dni = digits_only(row[2] if len(row) > 2 else None)
            if dni:
                if dni in seen_dni:
                    duplicate_dni.add(dni)
                else:
                    seen_dni.add(dni)

            birth_date = parse_date(row[3] if len(row) > 3 else None)
            parent_phone = digits_only(row[7] if len(row) > 7 else None)
            parent_name = clean_text(row[8] if len(row) > 8 else None)
            is_active = parse_active(row[1] if len(row) > 1 else None)

            player_record = {
                "source_key": f"{division_name}|{normalize_text(first_name)}|{normalize_text(last_name)}|{birth_date or ''}|{row_index}",
                "division_name": division_name,
                "division_id": divisions[division_name],
                "first_name": first_name or "",
                "last_name": last_name or "",
                "dni": dni,
                "birth_date": birth_date,
                "parent_phone": parent_phone,
                "parent_name": parent_name,
                "active": True,
                "inactivo": not is_active,
                "fecha_alta": today,
            }
            players.append(player_record)
            stats["players_by_division"][division_name] += 1

            for col_idx, session_date in date_columns:
                present = is_present(row[col_idx] if len(row) > col_idx else None)
                if present:
                    stats["present_cells"] += 1
                attendance_rows.append(
                    {
                        "source_key": player_record["source_key"],
                        "division_id": player_record["division_id"],
                        "session_date": session_date,
                        "present": present,
                    }
                )

    if duplicate_dni:
        for player in players:
            if player["dni"] in duplicate_dni:
                player["dni"] = None

    stats["dates"] = sorted(stats["dates"])
    return players, attendance_rows, stats


def insert_players(api: SupabaseRest, players: list[dict[str, Any]]) -> None:
    payload = [
        {
            "first_name": player["first_name"],
            "last_name": player["last_name"],
            "dni": player["dni"],
            "birth_date": player["birth_date"],
            "parent_phone": player["parent_phone"],
            "parent_name": player["parent_name"],
            "division_id": player["division_id"],
            "active": player["active"],
            "inactivo": player["inactivo"],
            "fecha_alta": player["fecha_alta"],
        }
        for player in players
    ]
    for batch in chunked(payload, 100):
        api.post("players", batch, "return=minimal")


def fetch_players_for_matching(api: SupabaseRest, division_ids: list[str]) -> list[dict[str, Any]]:
    return api.get(
        "players",
        {
            "select": "id,first_name,last_name,birth_date,division_id",
            "division_id": f"in.({','.join(division_ids)})",
        },
    )


def build_player_map(players: list[dict[str, Any]]) -> dict[str, str]:
    result: dict[str, str] = {}
    counts: dict[str, int] = defaultdict(int)
    for player in players:
        key = (
            f"{player['division_id']}|"
            f"{normalize_text(player.get('first_name'))}|"
            f"{normalize_text(player.get('last_name'))}|"
            f"{player.get('birth_date') or ''}"
        )
        counts[key] += 1
        result[key] = player["id"]
    ambiguous = {key for key, count in counts.items() if count > 1}
    for key in ambiguous:
        result.pop(key, None)
    return result


def ensure_sessions(api: SupabaseRest, division_ids: list[str], dates: list[str]) -> dict[tuple[str, str], str]:
    needed = [
        {"division_id": division_id, "session_date": date, "session_type": "sabado"}
        for division_id in division_ids
        for date in dates
    ]
    for batch in chunked(needed, 100):
        api.post("training_sessions", batch, "resolution=merge-duplicates,return=representation")

    quoted_dates = ",".join(f'"{date}"' for date in dates)
    rows = api.get(
        "training_sessions",
        {
            "select": "id,division_id,session_date",
            "division_id": f"in.({','.join(division_ids)})",
            "session_date": f"in.({quoted_dates})",
        },
    )
    return {(row["division_id"], row["session_date"]): row["id"] for row in rows}


def insert_attendance(
    api: SupabaseRest,
    attendance_rows: list[dict[str, Any]],
    player_key_map: dict[str, str],
    sessions: dict[tuple[str, str], str],
) -> None:
    payload: list[dict[str, Any]] = []
    for row in attendance_rows:
        division_id = row["division_id"]
        _, first_name, last_name, birth_date, _ = row["source_key"].split("|", 4)
        player_lookup_key = f"{division_id}|{first_name}|{last_name}|{birth_date}"
        player_id = player_key_map.get(player_lookup_key)
        if not player_id:
            raise ValueError(f"No pude matchear jugador para {row['source_key']}")
        session_id = sessions[(division_id, row["session_date"])]
        payload.append(
            {
                "session_id": session_id,
                "player_id": player_id,
                "present": row["present"],
            }
        )

    for batch in chunked(payload, 500):
        api.post("attendance_records", batch, "resolution=merge-duplicates,return=minimal")


def summarize(api: SupabaseRest, division_ids: list[str], dates: list[str]) -> list[dict[str, Any]]:
    quoted_dates = ",".join(f'"{date}"' for date in dates)
    sessions = api.get(
        "training_sessions",
        {
            "select": "id,session_date,division_id,divisions(name)",
            "division_id": f"in.({','.join(division_ids)})",
            "session_date": f"in.({quoted_dates})",
            "order": "session_date.asc",
        },
    )
    session_ids = [row["id"] for row in sessions]
    quoted_ids = ",".join(f'"{sid}"' for sid in session_ids)
    attendance = api.get(
        "attendance_records",
        {
            "select": "session_id,present",
            "session_id": f"in.({quoted_ids})",
        },
    )
    counts: dict[str, dict[str, int]] = defaultdict(lambda: {"present": 0, "total": 0})
    for row in attendance:
        counts[row["session_id"]]["total"] += 1
        if row["present"]:
            counts[row["session_id"]]["present"] += 1
    summary = []
    for session in sessions:
        division = session["divisions"]
        division_name = division["name"] if isinstance(division, dict) else division[0]["name"]
        summary.append(
            {
                "division": division_name,
                "date": session["session_date"],
                "present": counts[session["id"]]["present"],
                "total": counts[session["id"]]["total"],
            }
        )
    return summary


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", required=True, help="Ruta al Excel combinado")
    parser.add_argument("--execute", action="store_true", help="Inserta en Supabase")
    args = parser.parse_args()

    base_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    api_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not base_url or not api_key:
        print("Faltan NEXT_PUBLIC_SUPABASE_URL o SUPABASE_SERVICE_ROLE_KEY.", file=sys.stderr)
        return 1

    api = SupabaseRest(base_url, api_key)
    divisions = get_divisions(api)
    missing = [name for name in TARGET_DIVISIONS if name not in divisions]
    if missing:
        print(f"Faltan divisiones en Supabase: {', '.join(missing)}", file=sys.stderr)
        return 1

    players, attendance_rows, stats = build_plan(args.file, divisions)

    print("Vista previa de carga prerugby:")
    for division in TARGET_DIVISIONS:
        print(f"- {division}: {stats['players_by_division'][division]} jugadores")
    print(f"- Total jugadores: {len(players)}")
    print(f"- Fechas detectadas: {', '.join(stats['dates'])}")
    print(f"- Registros de asistencia: {len(attendance_rows)}")
    print(f"- Presentes marcados: {stats['present_cells']}")

    if not args.execute:
        print("")
        print("Modo vista previa. Usa --execute para importar.")
        return 0

    insert_players(api, players)
    division_ids = [divisions[name] for name in TARGET_DIVISIONS]
    inserted_players = fetch_players_for_matching(api, division_ids)
    player_key_map = build_player_map(inserted_players)
    sessions = ensure_sessions(api, division_ids, stats["dates"])
    insert_attendance(api, attendance_rows, player_key_map, sessions)

    print("")
    print("Resumen en Supabase:")
    for row in summarize(api, division_ids, stats["dates"]):
        print(f"- {row['division']} {row['date']}: {row['present']}/{row['total']}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
